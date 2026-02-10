#!/usr/bin/env python3
"""
财务报表注释Excel导出工具

将批量提取的财务报表注释导出为格式化的Excel文件
- Sheet 1: 目录（一级标题列表）
- Sheet 2-N: 各一级标题的详细内容
"""
import sys
import os
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any

# 添加项目根目录到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


class NotesExcelExporter:
    """财务报表注释Excel导出器"""

    def __init__(self):
        """初始化导出器"""
        # 颜色定义
        self.COLOR_HEADER = "1F4E78"  # 深蓝色
        self.COLOR_TOC_ALT = "F2F2F2"  # 浅灰色（目录斑马纹）
        self.COLOR_TITLE_BG = "D9E1F2"  # 浅蓝色（标题信息区）
        self.COLOR_CONTENT_HEADER = "70AD47"  # 深绿色（内容表头）
        self.COLOR_LEVEL1 = "B4C7E7"  # 浅蓝色（一级标题）
        self.COLOR_LEVEL2 = "FFF2CC"  # 浅黄色（二级标题）
        self.COLOR_TABLE_HEADER = "A6A6A6"  # 深灰色（表格表头）

        # 字体定义
        self.FONT_HEADER = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        self.FONT_TITLE = Font(name='微软雅黑', size=18, bold=True)
        self.FONT_NORMAL = Font(name='微软雅黑', size=11)
        self.FONT_BOLD = Font(name='微软雅黑', size=11, bold=True)

        # 边框定义
        thin_border = Side(style='thin', color='000000')
        self.BORDER_ALL = Border(left=thin_border, right=thin_border,
                                 top=thin_border, bottom=thin_border)

    def sanitize_sheet_name(self, name: str, max_length: int = 31) -> str:
        """
        清理sheet名称，符合Excel规范

        Args:
            name: 原始名称
            max_length: 最大长度（Excel限制31字符）

        Returns:
            清理后的名称
        """
        # 移除Excel不允许的字符
        name = re.sub(r'[\\/*?:\[\]]', '', name)

        # 移除序号前缀（如"1、 "）
        name = re.sub(r'^\d+、\s*', '', name)

        # 如果超长，截断并添加省略号
        if len(name) > max_length - 3:
            name = name[:max_length-3] + '...'

        return name

    def create_toc_sheet(self, wb: Workbook, notes_data: Dict[str, Any]) -> None:
        """
        创建目录sheet

        Args:
            wb: 工作簿对象
            notes_data: 注释数据
        """
        ws = wb.active
        ws.title = "目录"

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20

        # 表头
        headers = ['序号', '标题', '页码', '子项数量', '表格数量', '工作表名称']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER_ALL

        # 筛选一级标题
        level1_notes = [n for n in notes_data['notes'] if n['level'] == 1]

        # 填充数据
        for idx, note in enumerate(level1_notes, 1):
            row = idx + 1

            # 计算子项数量
            note_number = note['number']
            sub_notes = [n for n in notes_data['notes']
                        if n['level'] == 2 and n['number'].startswith(note_number + '.')]
            sub_count = len(sub_notes)

            # 计算表格数量
            table_count = note.get('content', {}).get('table_count', 0)
            for sub_note in sub_notes:
                table_count += sub_note.get('content', {}).get('table_count', 0)

            # 生成sheet名称
            sheet_name = f"{note['number']}_{self.sanitize_sheet_name(note['title'])}"

            # 填充单元格
            data = [
                note['number'],
                note['full_title'],
                note['page_num'],
                sub_count,
                table_count,
                sheet_name
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.FONT_NORMAL
                cell.alignment = Alignment(horizontal='left' if col == 2 else 'center',
                                         vertical='center')
                cell.border = self.BORDER_ALL

                # 斑马纹
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color=self.COLOR_TOC_ALT,
                                          end_color=self.COLOR_TOC_ALT,
                                          fill_type='solid')

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 添加自动筛选
        ws.auto_filter.ref = f"A1:F{len(level1_notes) + 1}"

    def create_content_sheet(self, wb: Workbook, note: Dict[str, Any],
                            all_notes: List[Dict[str, Any]]) -> None:
        """
        创建内容sheet

        Args:
            wb: 工作簿对象
            note: 一级标题注释数据
            all_notes: 所有注释数据（用于查找子项）
        """
        # 创建sheet
        sheet_name = f"{note['number']}_{self.sanitize_sheet_name(note['title'])}"
        ws = wb.create_sheet(title=sheet_name)

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 80

        current_row = 1

        # 区域1: 标题信息区
        current_row = self._write_title_section(ws, note, current_row)

        # 区域2: 内容区表头
        current_row += 1
        headers = ['层级', '标题', '页码', '内容']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_CONTENT_HEADER,
                                   end_color=self.COLOR_CONTENT_HEADER,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER_ALL

        current_row += 1

        # 写入一级标题内容
        current_row = self._write_note_content(ws, note, current_row, level=1)

        # 查找并写入子项
        # 由于LLM提取的编号可能不规范，使用页码和位置来判断父子关系
        # 查找同一页面上、在当前注释之后、在下一个一级标题之前的二级标题
        current_page = note['page_num']
        current_index = None

        # 找到当前注释在列表中的索引
        level1_notes = [n for n in all_notes if n['level'] == 1]
        for i, n in enumerate(level1_notes):
            if n['number'] == note['number'] and n['page_num'] == note['page_num']:
                current_index = i
                break

        # 找到下一个一级标题（可能在下一页）
        next_level1_page = None
        if current_index is not None and current_index < len(level1_notes) - 1:
            next_level1_page = level1_notes[current_index + 1]['page_num']

        # 查找属于当前一级标题的二级子项
        sub_notes = []
        for n in all_notes:
            if n['level'] == 2:
                # 二级标题必须在当前一级标题的页面或之后
                if n['page_num'] < current_page:
                    continue
                # 如果有下一个一级标题，二级标题必须在它之前
                if next_level1_page is not None and n['page_num'] >= next_level1_page:
                    continue
                # 如果在同一页，检查是否在当前一级标题之后
                # (这里假设notes列表已经按页面位置排序)
                if n['page_num'] == current_page:
                    # 找到当前注释和二级标题在all_notes中的位置
                    try:
                        note_pos = all_notes.index(note)
                        sub_pos = all_notes.index(n)
                        if sub_pos > note_pos:
                            sub_notes.append(n)
                    except ValueError:
                        pass
                elif n['page_num'] > current_page and (next_level1_page is None or n['page_num'] < next_level1_page):
                    sub_notes.append(n)

        for sub_note in sub_notes:
            current_row = self._write_note_content(ws, sub_note, current_row, level=2)

        # 冻结表头
        ws.freeze_panes = f'A{current_row - len(sub_notes) - 1}'

    def _write_title_section(self, ws, note: Dict[str, Any], start_row: int) -> int:
        """
        写入标题信息区

        Args:
            ws: 工作表对象
            note: 注释数据
            start_row: 起始行

        Returns:
            下一行行号
        """
        # 合并单元格 A1:D5
        ws.merge_cells(f'A{start_row}:D{start_row+4}')

        # 构建标题信息文本
        table_count = note.get('content', {}).get('table_count', 0)
        level_text = "一级标题" if note['level'] == 1 else "二级标题"

        info_text = (
            f"注释标题：{note['full_title']}\n"
            f"页码：{note['page_num']}\n"
            f"层级：{level_text}\n"
            f"包含表格：{table_count}个"
        )

        cell = ws[f'A{start_row}']
        cell.value = info_text
        cell.font = Font(name='微软雅黑', size=14, bold=True)
        cell.fill = PatternFill(start_color=self.COLOR_TITLE_BG,
                               end_color=self.COLOR_TITLE_BG,
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = self.BORDER_ALL

        return start_row + 5

    def _write_note_content(self, ws, note: Dict[str, Any],
                           start_row: int, level: int) -> int:
        """
        写入注释内容

        Args:
            ws: 工作表对象
            note: 注释数据
            start_row: 起始行
            level: 层级（1或2）

        Returns:
            下一行行号
        """
        content = note.get('content', {})
        text = content.get('text', '[无文本内容]')
        tables = content.get('tables', [])

        # 层级文本
        level_text = "一级" if level == 1 else "二级"

        # 写入标题行
        row_data = [level_text, note['full_title'], note['page_num'], text]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col, value=value)
            cell.font = self.FONT_BOLD if col <= 2 else self.FONT_NORMAL
            cell.alignment = Alignment(
                horizontal='center' if col <= 3 else 'left',
                vertical='top',
                wrap_text=True
            )
            cell.border = self.BORDER_ALL

            # 根据层级设置背景色
            if level == 1:
                cell.fill = PatternFill(start_color=self.COLOR_LEVEL1,
                                       end_color=self.COLOR_LEVEL1,
                                       fill_type='solid')
            elif level == 2:
                cell.fill = PatternFill(start_color=self.COLOR_LEVEL2,
                                       end_color=self.COLOR_LEVEL2,
                                       fill_type='solid')

        current_row = start_row + 1

        # 写入表格
        if tables:
            for table_idx, table in enumerate(tables, 1):
                current_row = self._write_table(ws, table, current_row, table_idx)
                current_row += 2  # 表格之间空2行

        return current_row

    def _write_table(self, ws, table: List[List], start_row: int,
                    table_idx: int) -> int:
        """
        写入表格

        Args:
            ws: 工作表对象
            table: 表格数据（二维列表）
            start_row: 起始行
            table_idx: 表格序号

        Returns:
            下一行行号
        """
        if not table:
            return start_row

        # 表格标题
        ws.merge_cells(f'A{start_row}:D{start_row}')
        cell = ws[f'A{start_row}']
        cell.value = f"表格 {table_idx}"
        cell.font = self.FONT_BOLD
        cell.fill = PatternFill(start_color=self.COLOR_TOC_ALT,
                               end_color=self.COLOR_TOC_ALT,
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.BORDER_ALL

        current_row = start_row + 1

        # 写入表格数据
        for row_idx, row_data in enumerate(table):
            for col_idx, cell_value in enumerate(row_data):
                # 表格从B列开始，留出A列空间
                col = col_idx + 2
                cell = ws.cell(row=current_row, column=col, value=cell_value)

                # 表头行样式
                if row_idx == 0:
                    cell.font = Font(name='微软雅黑', size=10, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color=self.COLOR_TABLE_HEADER,
                                          end_color=self.COLOR_TABLE_HEADER,
                                          fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(name='微软雅黑', size=10)
                    cell.alignment = Alignment(horizontal='right' if self._is_numeric(cell_value)
                                             else 'left',
                                             vertical='center')

                cell.border = self.BORDER_ALL

            current_row += 1

        return current_row

    def _is_numeric(self, value) -> bool:
        """判断值是否为数值"""
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            # 移除千分位逗号后尝试转换
            try:
                float(value.replace(',', ''))
                return True
            except (ValueError, AttributeError):
                return False
        return False

    def export_to_excel(self, json_file: str, output_file: str,
                       company_name: str = None) -> None:
        """
        导出注释到Excel

        Args:
            json_file: 输入的JSON文件路径
            output_file: 输出的Excel文件路径
            company_name: 公司名称（可选）
        """
        print(f"正在读取注释数据: {json_file}")

        # 读取JSON数据
        with open(json_file, 'r', encoding='utf-8') as f:
            notes_data = json.load(f)

        if not notes_data.get('success'):
            print("错误: 注释提取未成功")
            return

        total_notes = notes_data.get('total_notes', 0)
        print(f"找到 {total_notes} 个注释")

        # 创建工作簿
        wb = Workbook()

        # 创建目录sheet
        print("创建目录...")
        self.create_toc_sheet(wb, notes_data)

        # 创建内容sheets
        level1_notes = [n for n in notes_data['notes'] if n['level'] == 1]
        print(f"创建 {len(level1_notes)} 个内容工作表...")

        for idx, note in enumerate(level1_notes, 1):
            print(f"  [{idx}/{len(level1_notes)}] {note['full_title']}")
            self.create_content_sheet(wb, note, notes_data['notes'])

        # 设置文件元数据
        wb.properties.title = f"{company_name or '公司'} 财务报表注释"
        wb.properties.subject = "财务报表附注"
        wb.properties.creator = "PDF财务报表数据提取工具 v1.4.0"
        wb.properties.created = datetime.now()

        # 保存文件
        print(f"保存Excel文件: {output_file}")
        wb.save(output_file)
        print("✓ 导出完成！")


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='将财务报表注释导出为Excel文件')
    parser.add_argument('json_file', help='输入的JSON文件路径')
    parser.add_argument('-o', '--output', help='输出的Excel文件路径')
    parser.add_argument('-c', '--company', help='公司名称')

    args = parser.parse_args()

    # 设置默认输出路径
    if args.output is None:
        json_path = Path(args.json_file)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        company = args.company or '公司'
        output_name = f"{company}_财务报表注释_{timestamp}.xlsx"
        output_dir = Path('output')
        output_dir.mkdir(exist_ok=True)
        args.output = str(output_dir / output_name)

    # 执行导出
    exporter = NotesExcelExporter()
    exporter.export_to_excel(args.json_file, args.output, args.company)


if __name__ == '__main__':
    main()

